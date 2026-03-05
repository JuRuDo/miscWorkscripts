#!/bin/env python

import argparse
import openpyxl
import os
import math


def main():
    parser = argparse.ArgumentParser(description="Get length of all transcripts")
    parser.add_argument("-i", "--input", type=str, default=None, required=True,
                        help="path to input directory with csv files")
    parser.add_argument("-o", "--output", type=str, default=None, required=True,
                        help="output file")
    args = parser.parse_args()

    filelist = os.listdir(args.input)
    files = []
    for i in filelist:
        if i.endswith('.csv'):
            files.append(i)
    wb = openpyxl.Workbook()
    openpyxl.styles.DEFAULT_FONT.name = "Liberation Sans"
    f = 'LogFC'
    for i in files:
        wb.create_sheet(i.strip('.csv'))
    wb.remove(wb.worksheets[0])
    for i in range(len(files)):
        ws = wb.worksheets[i]
        with open(args.input + '/' + files[i], 'r') as infile:
            for line in infile.readlines():
                cells = line.rstrip('\n').split(',')
                cells.insert(1, cells.pop(5))
                cells.insert(2, cells.pop(6))
                if cells[3] == '"self.average"':
                    cells.insert(1, f)
                    for x in range(0, len(cells)):
                        cells[x] = cells[x].strip('"')
                        if cells[x].isnumeric():
                            cells[x] = float(cells[x])
                    ws.append(cells)
                else:
                    if not (cells[0].startswith('"Unassigned') or cells[0].startswith('"NegControl')):
                        try:
                            x = str(math.log2(float(cells[3])/float(cells[4])))
                        except:
                            x = ''
                        cells.insert(1, x)
                        for x in range(0, len(cells)):
                            cells[x] = cells[x].strip('"')
                            try:
                                cells[x] = float(cells[x])
                            except:
                                None
                        ws.append(cells)
            for column_cells in ws.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    ws.column_dimensions[new_column_letter].width = new_column_length * 1.1
            infile.close()
    wb.save(args.output)


if __name__ == '__main__':
    main()
