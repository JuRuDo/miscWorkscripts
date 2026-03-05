#!/bin/env python

import argparse
import openpyxl


def main():
    parser = argparse.ArgumentParser(description="Get length of all transcripts")
    parser.add_argument("-i", "--input", type=str, default=None, required=True,
                        help="path to input clsx file")
    parser.add_argument("-g", "--genes", type=str, default=None, required=True,
                        help="path to input file with list of genes")
    parser.add_argument("-o", "--output", type=str, default=None, required=True,
                        help="output csv file")
    args = parser.parse_args()

    geneList = loadGeneList(args.genes)
    wb = openpyxl.load_workbook(args.input, read_only=True)
    mutations = {}
    for i in range(len(wb.sheetnames)):
        ws = wb.worksheets[i]
        mutations[wb.sheetnames[i]] = {}
        tmp = []
        for row in ws.rows:
            tmp.append(row[0].value)
        for gene in geneList:
            if gene in tmp:
                mutations[wb.sheetnames[i]][gene] = 1
            else:
                mutations[wb.sheetnames[i]][gene] = 0
    writeMutations(args.output, mutations)


def writeMutations(path, mutations):
    with open(path, 'w') as outfile:
        line = '\t' + '\t'.join(mutations[list(mutations.keys())[0]]) + '\tAll\n'
        outfile.write(line)
        for sample in mutations.keys():
            outfile.write(sample)
            tmp = 0
            for gene in list(mutations[list(mutations.keys())[0]].keys()):
                outfile.write('\t' + str(mutations[sample][gene]))
                tmp += mutations[sample][gene]
            if tmp > 0:
                outfile.write('\t1\n')
            else:
                outfile.write('\t0\n')


def loadGeneList(path):
    geneList = []
    with open(path, 'r') as infile:
        for line in infile.readlines():
            geneList.append(line.strip('\n'))
    infile.close()
    return geneList
